# scripts/gen-gantt.py
#
# Builds HFSE-SIS-Implementation-Timeline.xlsx at the repo root.
#
#   Sheet "Timeline"    the project timeline table, styled to match the school's
#                       existing sheet: navy title band, blue phase bands, green
#                       status chips, navy footer.
#   Sheet "Gantt Chart" the same 29 activities as a stacked-bar Gantt.
#   Sheet "Chart Data"  the series the chart reads - an invisible offset sized to
#                       each start date, plus four duration series split by status
#                       so each gets its own colour.
#
# Why Python: the repo's `xlsx` package is the SheetJS community build and can
# write neither cell fills nor chart objects. openpyxl does both.
#
# STRICTLY READ-ONLY with respect to the database. Writes exactly one file.
#
# Run:
#   python scripts/gen-gantt.py

import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("Needs openpyxl:  pip install openpyxl\n")
    sys.exit(1)

# palette lifted from the school's existing timeline sheet
NAVY = "12284C"
BAND = "2E75B6"
PHASE = "D9E2F3"
CYAN = "00B0F0"
WHITE = "FFFFFF"
GREEN = "00B050"
AMBER = "FFC000"
SKY = "41A5EE"
GREY = "A6A6A6"
INK = "1F3864"
LINE = "BFBFBF"

EPOCH = date(1899, 12, 30)
serial = lambda d: (d - EPOCH).days
AXIS_MIN, AXIS_MAX = serial(date(2026, 3, 16)), serial(date(2026, 10, 26))

MON = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
D = date

# (no, phase, activity, start, end, duration, status, who, notes)
ROWS = [
 (1,"PHASE 1: KICKOFF","Project Kickoff & Scope Confirmation",D(2026,3,17),D(2026,3,22),"6 days","Finished","Amier / Ace","Charter, RACI, comms cadence"),
 (2,"PHASE 2: DISCOVERY","User Research & Requirements Gathering",D(2026,3,23),D(2026,3,29),"7 days","Finished","Amier / Ace","Interviews with grading coordinator"),
 (3,"PHASE 3: FOUNDATION","SIS Admin Module",D(2026,4,20),D(2026,7,14),"~12 weeks","Finished","VizBytes Engineering","Year setup, roles, calendar, subjects. Navigation redesign in July"),
 (4,"PHASE 4: CORE BUILD","Markbook (Grading) Module",D(2026,4,14),D(2026,6,6),"~8 weeks","Finished","VizBytes Engineering","Server-side computation, audit log, Masterfile export"),
 (5,"PHASE 4: CORE BUILD","Records Module",D(2026,4,17),D(2026,6,12),"~8 weeks","Finished","VizBytes Engineering","Student profile, family, enrolment"),
 (6,"PHASE 4: CORE BUILD","Admissions Module",D(2026,4,17),D(2026,8,5),"~16 weeks","Finished","VizBytes Engineering","Application pipeline and document workflow"),
 (7,"PHASE 4: CORE BUILD","P-Files Module",D(2026,4,17),D(2026,7,30),"~15 weeks","Finished","VizBytes Engineering","Document repository with versioning"),
 (8,"PHASE 5: INTEGRATION","Attendance Module",D(2026,4,21),D(2026,6,26),"~10 weeks","Finished","VizBytes Engineering","Daily register feeding the report card"),
 (9,"PHASE 5: INTEGRATION","Student Evaluation Module",D(2026,4,22),D(2026,5,30),"~6 weeks","Finished","VizBytes Engineering","Virtue theme and adviser write-ups"),
 (10,"PHASE 5: INTEGRATION","Cross-Module Integration & QA",D(2026,4,14),D(2026,8,24),"~19 weeks","Finished","VizBytes Engineering","Continuous, not a fixed window. 3,152 automated checks green"),
 (11,"PHASE 6: DEMO PREP","Demo Environment & Test Data",D(2026,5,5),D(2026,7,30),"~12 weeks","Finished (retired)","Ace","Retired 30 Jul once real school data was loaded"),
 (12,"PHASE 6: DEMO PREP","Executive Demo Presentation",D(2026,5,14),D(2026,5,18),"5 days","Finished","Amier / Ace","Deck and walkthrough script"),
 (13,"PHASE 7: DEVELOPMENT OF REVISIONS","Post-Demo Revisions & Refinement",D(2026,5,18),D(2026,8,24),"~14 weeks","Finished","VizBytes Engineering","CORRECTED - ran to 24 Aug, not 5 Jul. Detail in rows 15-17"),
 (14,"PHASE 7: DEVELOPMENT OF REVISIONS","Historical Data Migration",D(2026,7,9),D(2026,7,20),"12 days","Finished","VizBytes Engineering","AY2025 T1-T3 and AY2026 T1-T2 loaded. AY2025 T4 blocked - no source files"),
 (15,"PHASE 7b: EXTENDED BUILD","Platform & Workflow Extensions",D(2026,4,15),D(2026,7,31),"~15 weeks","Finished","VizBytes Engineering","ADDED. Change requests, dashboards, parent portal, test pipeline, calendar, insights, year setup, grade levels, subject setup, permissions"),
 (16,"PHASE 7b: EXTENDED BUILD","Classroom Module & Teacher Tools",D(2026,7,29),D(2026,8,21),"~3 weeks","Finished","VizBytes Engineering","ADDED. Per-class workspace: health, timeline, student details, at-risk list, discipline filing"),
 (17,"PHASE 7b: EXTENDED BUILD","Post-Training Enhancements",D(2026,8,3),D(2026,8,17),"~2 weeks","Finished","VizBytes Engineering","ADDED. Excused-absence reason, House colours, whole-year trend, save/confirm across all 80 screens"),
 (18,"PHASE 8: FINAL PRESENTATION","Final System Presentation to Leadership",D(2026,7,15),D(2026,7,19),"5 days","Finished","Amier / Ace","Go-live readiness review"),
 (19,"PHASE 9: VIRTUAL TRAINING","Travel - MNL to SIN",None,None,"-","CANCELLED","Ace","Training moved to virtual"),
 (20,"PHASE 9: VIRTUAL TRAINING","Virtual Faculty / Teacher Training",D(2026,7,31),D(2026,7,31),"1 day","Finished","Ace","Held 2 weeks EARLY (planned 14 Aug). 11 feedback items raised"),
 (21,"PHASE 9: VIRTUAL TRAINING","Virtual Admin & Registrar Training",D(2026,8,13),D(2026,8,13),"1 day","Finished","Ace","Held early (planned 18 Aug). 13 action items raised"),
 (22,"PHASE 9: VIRTUAL TRAINING","Virtual Power Users / Champions Training",None,None,"-","Not yet held","Ace","Original date 19 Aug has passed. Needs a new date"),
 (23,"PHASE 9: VIRTUAL TRAINING","Wrap-up & Handover",D(2026,8,30),D(2026,8,30),"1 day","Upcoming","Ace","Training materials handover"),
 (24,"POST-LAUNCH","Consolidation of Additional Features for 2nd Phase Dev",D(2026,8,18),D(2026,10,26),"~10 weeks","In progress","Amier / Ace","Phase 2 backlog - see rows 28-29"),
 (25,"POST-LAUNCH","System Implementation (Parallel with Current System)",D(2026,8,13),D(2026,10,26),"~11 weeks","In progress","Amier / Ace","Admin team committed 13 Aug to run alongside the current panel"),
 (26,"POST-LAUNCH","System Implementation (Full Transition)",None,None,"-","Not scheduled","Amier / Ace","Follows a clean parallel run. Date to be agreed with the school"),
 (27,"POST-LAUNCH","Platform Stabilisation & Monthly Support",D(2026,8,25),D(2026,10,26),"~9 weeks","Upcoming","Amier / Ace","Hypercare alongside the parallel run"),
 (28,"PHASE 2 IMPLEMENTATION","Phase 2 - Delivered (3 of 10)",D(2026,8,3),D(2026,8,24),"~3 weeks","SHIPPED","VizBytes Engineering","Disciplinary Records, House Colours, Relief Teacher Covers. Board still shows 0%"),
 (29,"PHASE 2 IMPLEMENTATION","Phase 2 - Awaiting School Input (7 of 10)",None,None,"-","Blocked","VizBytes Engineering","Awards, MC/Travel Declaration, P-Files list, Supplies, SOW, TOR, AEB approval flow"),
]

STATUS_FILL = {
    "Finished": GREEN, "SHIPPED": GREEN,
    "In progress": AMBER, "Upcoming": SKY,
    "Finished (retired)": GREY, "CANCELLED": GREY, "Withdrawn": GREY,
}
BUCKET = {"Finished": 0, "SHIPPED": 0, "In progress": 1,
          "Upcoming": 2, "Finished (retired)": 3, "CANCELLED": 3, "Withdrawn": 3}


def fmt_dates(s, e):
    if s is None:
        return "-"
    if s == e:
        return f"{MON[s.month-1]} {s.day}, {s.year}"
    if s.month == e.month:
        return f"{MON[s.month-1]} {s.day}-{e.day}, {s.year}"
    return f"{MON[s.month-1]} {s.day} - {MON[e.month-1]} {e.day}, {e.year}"


def thin(c=LINE):
    return Side(style="thin", color=c)


def box(c=LINE):
    return Border(left=thin(c), right=thin(c), top=thin(c), bottom=thin(c))


def build_timeline(ws):
    LAST = 8  # A spacer + B..H
    widths = {1: 2, 2: 5, 3: 46, 4: 22, 5: 12, 6: 18, 7: 21, 8: 62}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    def fill_row(r, colour, height=None):
        for c in range(2, LAST + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=colour)
        if height:
            ws.row_dimensions[r].height = height

    # title band
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=LAST)
    t = ws.cell(row=2, column=2, value="IMPLEMENTATION TIMELINE  |  HFSE INTERNATIONAL SCHOOL")
    t.font = Font(name="Calibri", size=16, bold=True, color=CYAN)
    t.alignment = Alignment(horizontal="center", vertical="center")
    fill_row(2, NAVY, 32)

    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=LAST)
    s = ws.cell(row=3, column=2,
                value="Student Information System (SIS) - Actual Delivery Against Plan "
                      "(Mar 17 - Oct 26, 2026)")
    s.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    fill_row(3, NAVY, 22)

    ws.row_dimensions[4].height = 8

    # PROJECT TIMELINE band
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=LAST)
    p = ws.cell(row=5, column=2, value="PROJECT TIMELINE")
    p.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    p.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    fill_row(5, BAND, 20)

    # header
    for c, h in enumerate(["No.", "Activity", "Dates", "Duration",
                           "Status", "Responsible", "Notes"], start=2):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box(WHITE)
    ws.row_dimensions[6].height = 24

    r = 7
    seen = None
    for no, phase, act, st, en, dur, status, who, notes in ROWS:
        if phase != seen:
            seen = phase
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=LAST)
            b = ws.cell(row=r, column=2, value=phase)
            b.font = Font(name="Calibri", size=10, bold=True, color=INK)
            b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            fill_row(r, PHASE, 18)
            r += 1

        vals = [no, act, fmt_dates(st, en), dur, status, who, notes]
        for c, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=WHITE)
            cell.border = box()
            if c == 2:
                cell.font = Font(name="Calibri", size=9, color="808080")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 3:
                cell.font = Font(name="Calibri", size=10, bold=True, color=INK)
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True, indent=1)
            elif c == 6:
                f = STATUS_FILL.get(status)
                if f:
                    cell.fill = PatternFill("solid", fgColor=f)
                cell.font = Font(name="Calibri", size=9, bold=True,
                                 color=WHITE if f == GREY else "000000")
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
            elif c == 8:
                cell.font = Font(name="Calibri", size=8.5, color="404040")
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True, indent=1)
            else:
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
        ws.row_dimensions[r].height = 30
        r += 1

    # footer
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=LAST)
    f = ws.cell(row=r, column=2,
                value="Total Implementation Period: March 17 - August 24, 2026 "
                      "(~23 weeks, 10 phases)  +  post-launch support "
                      "(August 25 - October 26, 2026)")
    f.font = Font(name="Calibri", size=11, bold=True, color=CYAN)
    f.alignment = Alignment(horizontal="center", vertical="center")
    fill_row(r, NAVY, 26)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B7"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_data(ws):
    ws.append(["Activity", "Offset", "Delivered", "In progress", "Upcoming", "Withdrawn"])
    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
    for i, (no, phase, act, st, en, dur, status, who, notes) in enumerate(ROWS):
        r = 2 + i
        ws.cell(row=r, column=1, value=act)
        if st and en:
            ws.cell(row=r, column=2, value=serial(st))
            b = BUCKET.get(status)
            days = (en - st).days + 1
            for k in range(4):
                ws.cell(row=r, column=3 + k, value=days if k == b else 0)
    ws.column_dimensions["A"].width = 46
    for col in "BCDEF":
        ws.column_dimensions[col].width = 11
    ws.sheet_view.showGridLines = False


def build_chart(ws_chart, ws_data):
    ch = BarChart()
    ch.type = "bar"
    ch.grouping = "stacked"
    ch.overlap = 100
    ch.gapWidth = 30
    ch.title = "HFSE SIS - Implementation Timeline (Mar - Oct 2026)"

    last = 1 + len(ROWS)
    ch.add_data(Reference(ws_data, min_col=2, max_col=6, min_row=1, max_row=last),
                titles_from_data=True)
    ch.set_categories(Reference(ws_data, min_col=1, min_row=2, max_row=last))

    spacer = ch.series[0]
    spacer.graphicalProperties = GraphicalProperties()
    spacer.graphicalProperties.noFill = True
    spacer.graphicalProperties.line = LineProperties(noFill=True)

    for i, colour in enumerate([GREEN, AMBER, SKY, GREY], start=1):
        s = ch.series[i]
        s.graphicalProperties = GraphicalProperties(solidFill=colour)
        s.graphicalProperties.line = LineProperties(solidFill=colour)

    ch.y_axis.scaling.min = AXIS_MIN
    ch.y_axis.scaling.max = AXIS_MAX
    ch.y_axis.numFmt = "d mmm"
    ch.y_axis.delete = False
    ch.x_axis.scaling.orientation = "maxMin"
    ch.x_axis.delete = False

    ch.height = 30
    ch.width = 34
    ws_chart.add_chart(ch, "B2")
    ws_chart.sheet_view.showGridLines = False


def main():
    out = Path(__file__).resolve().parent.parent / "HFSE-SIS-Implementation-Timeline.xlsx"
    wb = Workbook()
    tl = wb.active
    tl.title = "Timeline"
    gc = wb.create_sheet("Gantt Chart")
    cd = wb.create_sheet("Chart Data")

    build_timeline(tl)
    build_data(cd)
    build_chart(gc, cd)

    try:
        wb.save(out)
    except PermissionError:
        sys.stderr.write(f"{out.name} is open in Excel. Close it and re-run.\n")
        return 1

    bars = sum(1 for r in ROWS if r[3])
    phases = len({r[1] for r in ROWS})
    print(f"Wrote {out}")
    print(f"  Timeline:  {len(ROWS)} activities across {phases} phase bands")
    print(f"  Gantt:     {bars} bars, {len(ROWS) - bars} undated rows")
    return 0


if __name__ == "__main__":
    sys.exit(main())
