# scripts/gen-project-timeline.py
#
# Generates HFSE-SIS-Implementation-Timeline.xlsx at the repo root: the project
# implementation timeline rendered as a weekly-column Gantt chart, covering the
# original March 2026 plan, what was actually delivered against it, and the
# Phase 2 backlog carried forward.
#
# Why a Python script and not the repo's usual `npx tsx`:
# the only Excel library in package.json is `xlsx@0.18.5`, the SheetJS COMMUNITY
# build, which cannot write cell fills, fonts or borders (the repo already notes
# this in lib/sis/provisioning/credential-workbook.ts). A Gantt chart is nothing
# but coloured cells, so that library physically cannot produce this file.
# openpyxl can, it is already installed, and going this way keeps a one-off
# management document out of the Next.js dependency tree entirely.
# scripts/md-to-docx.py is the existing precedent for a Python document generator
# in this repo.
#
# The row data below is a literal table, deliberately. This document is meant to
# be re-generated as the project moves: edit ROWS / DELIVERY_LOG and re-run.
#
# STRICTLY READ-ONLY with respect to the database. It touches no Supabase client,
# reads no student data, and writes exactly one file. The output contains no PII
# and is gitignored by the root `/*.xlsx` rule.
#
# Run:
#   python scripts/gen-project-timeline.py
#
# Exit code 0 on success, 1 if openpyxl is missing or the file cannot be written.

import sys
from datetime import date, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "openpyxl is required. Install it with:  pip install openpyxl\n"
    )
    sys.exit(1)


# --------------------------------------------------------------------------
# Palette - carried over from the existing timeline sheet so the regenerated
# document reads as the same artefact, not a different one.
# --------------------------------------------------------------------------

NAVY = "1F3864"
BLUE = "2E75B6"
CYAN_TEXT = "00B0F0"
WHITE = "FFFFFF"
BLACK = "000000"

GREEN = "00B050"          # delivered
AMBER = "FFC000"          # in progress / not yet verified
GREY = "A6A6A6"           # withdrawn or cancelled
LIGHTBLUE = "BDD7EE"      # proposed / scheduled ahead
RED = "C00000"            # today marker

ROW_ALT = "F2F7FB"        # zebra striping on the left block
GRIDLINE = "D9D9D9"

BAR_FILLS = {
    "done": GREEN,
    "wip": AMBER,
    "dropped": GREY,
    "proposed": LIGHTBLUE,
}

STATUS_FILLS = {
    "Finished": GREEN,
    "Finished (retired)": GREY,
    "Withdrawn": GREY,
    "CANCELLED": GREY,
    "Ongoing": AMBER,
    "In progress": AMBER,
    "Delivered - unverified": AMBER,
    "SHIPPED": GREEN,
}


# --------------------------------------------------------------------------
# Week grid
# --------------------------------------------------------------------------

GRID_START = date(2026, 3, 16)   # Monday before kickoff
GRID_END = date(2026, 10, 26)    # Monday, end of the proposed hypercare window

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def week_starts(start, end):
    """Every Monday from start to end inclusive."""
    weeks, cur = [], start
    while cur <= end:
        weeks.append(cur)
        cur += timedelta(days=7)
    return weeks


WEEKS = week_starts(GRID_START, GRID_END)
FIRST_WEEK_COL = 9  # column I - the left block occupies A..H


def week_index(d):
    """Index of the week containing date d, clamped to the grid."""
    if d < GRID_START:
        return 0
    if d > GRID_END + timedelta(days=6):
        return len(WEEKS) - 1
    return (d - GRID_START).days // 7


# --------------------------------------------------------------------------
# Date / duration formatting
# --------------------------------------------------------------------------

def fmt_dates(start, end):
    if start is None:
        return "To be scheduled"
    if start == end:
        return f"{start.day} {MONTHS[start.month - 1]} {start.year}"
    if start.month == end.month and start.year == end.year:
        return f"{start.day}-{end.day} {MONTHS[start.month - 1]} {start.year}"
    return (f"{start.day} {MONTHS[start.month - 1]} - "
            f"{end.day} {MONTHS[end.month - 1]} {end.year}")


def fmt_duration(start, end):
    if start is None:
        return "-"
    days = (end - start).days + 1
    if days <= 13:
        return "1 day" if days == 1 else f"{days} days"
    weeks = round(days / 7)
    return f"~{weeks} weeks"


# --------------------------------------------------------------------------
# The timeline. Each entry is either a phase band or an activity row.
#
#   band     - a phase heading
#   act      - (activity, start, end, status, responsible, change, notes, bar)
#              start/end may be None for work that has no agreed date.
#              `dates` overrides the auto-formatted date text where the planned
#              and actual dates both need to be visible.
# --------------------------------------------------------------------------

D = date

ROWS = [
    ("band", "PHASE 1: KICKOFF"),
    ("act", dict(
        activity="Project Kickoff & Scope Confirmation",
        start=D(2026, 3, 17), end=D(2026, 3, 22),
        status="Finished", who="Amier / Ace", change="As planned",
        notes="Charter, RACI, comms cadence.",
        bar="done")),

    ("band", "PHASE 2: DISCOVERY"),
    ("act", dict(
        activity="User Research & Requirements Gathering",
        start=D(2026, 3, 23), end=D(2026, 3, 29),
        status="Finished", who="Amier / Ace", change="As planned",
        notes="Interviews with the grading coordinator.",
        bar="done")),

    ("band", "PHASE 3: FOUNDATION"),
    ("act", dict(
        activity="SIS Admin Module",
        start=D(2026, 4, 20), end=D(2026, 7, 14),
        status="Finished", who="VizBytes Engineering", change="Date corrected",
        notes="Year setup, roles, calendar, subjects. Hub delivered 20 Apr; "
              "the full navigation redesign followed 11-14 Jul.",
        bar="done")),

    ("band", "PHASE 4: CORE BUILD"),
    ("act", dict(
        activity="Markbook (Grading) Module",
        start=D(2026, 4, 14), end=D(2026, 6, 6),
        status="Finished", who="VizBytes Engineering", change="Date corrected",
        notes="Server-side grade computation and audit log. Masterfile and the "
              "Excel report-book export followed 4 Jun.",
        bar="done")),
    ("act", dict(
        activity="Records Module",
        start=D(2026, 4, 17), end=D(2026, 6, 12),
        status="Finished", who="VizBytes Engineering", change="Date corrected",
        notes="Student profile, family, enrolment. The inquiry-tracking phase "
              "was dropped 24 Apr.",
        bar="done")),
    ("act", dict(
        activity="Admissions Module",
        start=D(2026, 4, 17), end=D(2026, 8, 5),
        status="Finished", who="VizBytes Engineering", change="Date corrected",
        notes="Application pipeline and document workflow. SharePoint inquiry "
              "sync dropped 24 Apr; enrolment and class assignment separated 5 Aug.",
        bar="done")),
    ("act", dict(
        activity="P-Files Module",
        start=D(2026, 4, 17), end=D(2026, 7, 30),
        status="Finished", who="VizBytes Engineering", change="Date corrected",
        notes="Document repository with versioning. Redesigned from a review "
              "queue to a repository on the day it was first built.",
        bar="done")),

    ("band", "PHASE 5: INTEGRATION"),
    ("act", dict(
        activity="Attendance Module",
        start=D(2026, 4, 21), end=D(2026, 6, 26),
        status="Finished", who="VizBytes Engineering", change="Date corrected",
        notes="Daily register feeding the report card. The Excel term-sheet "
              "replica landed 25-26 Jun.",
        bar="done")),
    ("act", dict(
        activity="Student Evaluation Module",
        start=D(2026, 4, 22), end=D(2026, 5, 30),
        status="Finished", who="VizBytes Engineering", change="Date corrected",
        notes="Virtue theme and adviser write-ups. Scope narrowed to adviser "
              "write-ups only on 30 May.",
        bar="done")),
    ("act", dict(
        activity="Cross-Module Integration & QA",
        start=D(2026, 4, 14), end=D(2026, 8, 24),
        status="Ongoing", who="VizBytes Engineering", change="Date corrected",
        notes="Continuous rather than a fixed window. 3,152 automated checks "
              "now run green on every change.",
        bar="done")),

    ("band", "PHASE 6: DEMO PREP"),
    ("act", dict(
        activity="Demo Environment & Test Data",
        start=D(2026, 5, 5), end=D(2026, 7, 30),
        status="Finished (retired)", who="Ace", change="Date corrected",
        notes="Realistic dataset for the demo. Retired 30 Jul once real school "
              "data was loaded and it was no longer needed.",
        bar="dropped")),
    ("act", dict(
        activity="Executive Demo Presentation",
        start=D(2026, 5, 14), end=D(2026, 5, 18),
        status="Finished", who="Amier / Ace", change="As planned",
        notes="Deck and walkthrough script.",
        bar="done")),

    ("band", "PHASE 7: DEVELOPMENT OF REVISIONS"),
    ("act", dict(
        activity="Post-Demo Revisions & Refinement",
        start=D(2026, 5, 18), end=D(2026, 8, 24),
        status="Finished", who="VizBytes Engineering", change="Date corrected",
        notes="The previous sheet closed this on 5 Jul. Revision work in fact "
              "ran continuously to 24 Aug - itemised in Phase 7b below.",
        bar="done")),
    ("act", dict(
        activity="Data Migration - AY2025 Historical Backfill",
        start=D(2026, 7, 9), end=D(2026, 7, 10),
        status="Finished", who="VizBytes Engineering", change="Date corrected",
        notes="AY2025 Term 1-3 grades recovered from the school's spreadsheets. "
              "Term 4 could not be loaded - no source workbooks exist.",
        bar="done")),
    ("act", dict(
        activity="Data Migration - AY2026 Term 1 & 2 Backfill",
        start=D(2026, 7, 17), end=D(2026, 7, 20),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="Enrolment, attendance, grading and evaluation records imported "
              "for both terms, Primary and Secondary.",
        bar="done")),

    ("band", "PHASE 7b: EXTENDED BUILD - delivered, not on the original plan"),
    ("act", dict(
        activity="Change Requests & Approval Workflow",
        start=D(2026, 4, 15), end=D(2026, 6, 5),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="Requests to edit a locked sheet, with live badges and one-click "
              "approve or reject straight from the email.",
        bar="done")),
    ("act", dict(
        activity="Dashboards & Drill-downs",
        start=D(2026, 4, 24), end=D(2026, 5, 6),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="A dashboard per module with drill-through to the underlying "
              "students.",
        bar="done")),
    ("act", dict(
        activity="Parent Portal - Sign-in & Report Card Access",
        start=D(2026, 4, 29), end=D(2026, 7, 30),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="Secure hand-off from the parent portal. Earlier-term adviser "
              "comments and the school letterhead added 30 Jul.",
        bar="done")),
    ("act", dict(
        activity="Automated Test Suite & Release Pipeline",
        start=D(2026, 5, 28), end=D(2026, 8, 24),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="Every change is now checked automatically before it ships.",
        bar="done")),
    ("act", dict(
        activity="Scheme of Work - built, then withdrawn",
        start=D(2026, 5, 20), end=D(2026, 5, 28),
        status="Withdrawn", who="VizBytes Engineering", change="Added",
        notes="Built and then removed: teachers keep their scheme of work in "
              "external documents. Reopened by Ms Christina on 21 Aug - see "
              "Phase 2 below.",
        bar="dropped")),
    ("act", dict(
        activity="School Calendar Rebuild",
        start=D(2026, 6, 4), end=D(2026, 6, 4),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="Month, week, day and list views; five day types; edit a single "
              "day without disturbing the pattern.",
        bar="done")),
    ("act", dict(
        activity="Insights Dashboards (four modules)",
        start=D(2026, 6, 10), end=D(2026, 7, 22),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="Enrolment Health, Retention & Population, Attendance Health and "
              "Academic Performance, each comparing two academic years.",
        bar="done")),
    ("act", dict(
        activity="Year Setup Workbench",
        start=D(2026, 6, 26), end=D(2026, 7, 8),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="A guided readiness checklist for rolling the school over into a "
              "new academic year.",
        bar="done")),
    ("act", dict(
        activity="Grade Levels & Progression",
        start=D(2026, 7, 11), end=D(2026, 7, 21),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="Levels became managed data with progression rules and alternate "
              "names, instead of being fixed in code.",
        bar="done")),
    ("act", dict(
        activity="Unified Subject Setup",
        start=D(2026, 7, 14), end=D(2026, 7, 16),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="One place to manage subjects and their weights, with per-class "
              "overrides.",
        bar="done")),
    ("act", dict(
        activity="Permissions & Roles Layer",
        start=D(2026, 7, 23), end=D(2026, 7, 31),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="What each role can do is now editable in the system rather than "
              "fixed in code. Registrar renamed to academic coordinator.",
        bar="done")),
    ("act", dict(
        activity="Home Page & Account Rebuild",
        start=D(2026, 7, 24), end=D(2026, 7, 25),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="A landing page that offers each person the work their job "
              "actually involves, with their to-dos and recent actions.",
        bar="done")),
    ("act", dict(
        activity="Classroom Module",
        start=D(2026, 7, 29), end=D(2026, 8, 21),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="A per-class workspace for teachers: class health, timeline, "
              "student details, at-risk list and discipline filing.",
        bar="done")),
    ("act", dict(
        activity="Training-Driven Features (Session 1 feedback)",
        start=D(2026, 8, 3), end=D(2026, 8, 9),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="Excused-absence reason, House colours, whole-year grade trend, "
              "and at-risk detection across all grade components.",
        bar="done")),
    ("act", dict(
        activity="Save & Confirm Feedback Across the App",
        start=D(2026, 8, 14), end=D(2026, 8, 17),
        status="Finished", who="VizBytes Engineering", change="Added",
        notes="All 80 save actions now confirm clearly and refresh the screen "
              "before telling the user they succeeded.",
        bar="done")),

    ("band", "PHASE 8: FINAL PRESENTATION"),
    ("act", dict(
        activity="Final System Presentation to Leadership",
        start=D(2026, 7, 15), end=D(2026, 7, 19),
        status="Finished", who="Amier / Ace", change="As planned",
        notes="Go-live readiness review.",
        bar="done")),

    ("band", "PHASE 9: VIRTUAL TRAINING"),
    ("act", dict(
        activity="Travel - MNL to SIN",
        start=None, end=None,
        status="CANCELLED", who="Ace", change="Cancelled",
        dates="22 Jun 2026 (cancelled)",
        notes="Arrive the day before training. Training moved to virtual.",
        bar=None)),
    ("act", dict(
        activity="Virtual Faculty / Teacher Training",
        start=D(2026, 7, 31), end=D(2026, 7, 31),
        status="Finished", who="Ace", change="Date corrected",
        dates="Held 31 Jul (planned 14 Aug)",
        notes="Held two weeks AHEAD of plan. Markbook, Attendance and "
              "Evaluation. Raised 11 feedback items, 5 already delivered.",
        bar="done")),
    ("act", dict(
        activity="Virtual Admin & Registrar Training",
        start=D(2026, 8, 13), end=D(2026, 8, 13),
        status="Finished", who="Ace", change="Date corrected",
        dates="Held 13 Aug (planned 18 Aug)",
        notes="Held ahead of plan, 1h 35m. SIS Admin, Records and P-Files. "
              "Raised 13 action items, 3 already delivered.",
        bar="done")),
    ("act", dict(
        activity="Virtual Power Users / Champions Training",
        start=None, end=None,
        status="Not yet held", who="Ace", change="Needs a date",
        dates="Planned 19 Aug - passed",
        notes="Train-the-trainer for HFSE. The original date has passed; a new "
              "date needs to be agreed.",
        bar=None)),
    ("act", dict(
        activity="Wrap-up & Handover",
        start=D(2026, 8, 30), end=D(2026, 8, 30),
        status="Upcoming", who="Ace", change="As planned",
        notes="Training materials handover.",
        bar="proposed")),

    ("band", "POST-LAUNCH"),
    ("act", dict(
        activity="Parallel Run with the Current System",
        start=D(2026, 8, 13), end=date.today(),
        status="In progress", who="Amier / Ace", change="Date corrected",
        notes="The admin team committed on 13 Aug to run the SIS alongside "
              "their current panel rather than cutting over.",
        bar="wip")),
    ("act", dict(
        activity="Hypercare & Monthly Support",
        start=D(2026, 8, 25), end=D(2026, 10, 26),
        status="Upcoming", who="Amier / Ace", change="Dated",
        notes="Support window running alongside the parallel run.",
        bar="proposed")),
    ("act", dict(
        activity="Full System Transition",
        start=None, end=None,
        status="Not scheduled", who="Amier / Ace", change="Needs a date",
        notes="Follows a clean parallel run. Date to be agreed with the school.",
        bar=None)),

    ("band", "PHASE 2 IMPLEMENTATION - backlog status as at 25 Aug 2026"),
    ("act", dict(
        activity="Student Disciplinary Records",
        start=D(2026, 8, 18), end=D(2026, 8, 21),
        status="SHIPPED", who="VizBytes Engineering", change="Board says 0%",
        notes="DELIVERED. Five screens plus a school-wide register; checked in "
              "the browser 24 Aug. The system records incidents - it decides "
              "nothing and generates no letters, by design.",
        bar="done")),
    ("act", dict(
        activity="House Colour & House-Points Tracking",
        start=D(2026, 8, 3), end=D(2026, 8, 6),
        status="SHIPPED", who="VizBytes Engineering", change="Board says 0%",
        notes="DELIVERED. Four houses named and student allocation loaded from "
              "Mr Hanafi's file. Confirm the 402-student import ran in "
              "production before reporting this 100% complete.",
        bar="done")),
    ("act", dict(
        activity="Relief Teacher Covers",
        start=D(2026, 8, 12), end=D(2026, 8, 24),
        status="Delivered - unverified", who="VizBytes Engineering",
        change="Board says 0%",
        notes="BUILT, with cover start and end dates added 24 Aug. Not yet "
              "checked in the browser, and only 3 teacher assignments exist to "
              "test against - needs the 26-teacher deployment to prove at scale.",
        bar="wip")),
    ("act", dict(
        activity="Awards - Certificates & Participation",
        start=None, end=None,
        status="Open", who="VizBytes Engineering", change="Awaiting the school",
        notes="Sample certificate received 14 Aug but not in the expected "
              "format. Awaiting related files from Ms Chandana and Ms Tin.",
        bar=None)),
    ("act", dict(
        activity="MC & Travel Declaration Upload",
        start=None, end=None,
        status="Open", who="VizBytes Engineering", change="Reshaped 17 Aug",
        notes="Now the PARENT files this through the portal, not the teacher - "
              "they hold the document. Approvers confirmed 19 Aug: form class "
              "adviser, then officer in charge. Larger than originally scoped.",
        bar=None)),
    ("act", dict(
        activity="Expand P-Files Document List",
        start=None, end=None,
        status="Blocked", who="VizBytes Engineering", change="Awaiting Ms Wynne",
        notes="Awaiting her list of documents to track for new and current "
              "students. Requested 21 Aug.",
        bar=None)),
    ("act", dict(
        activity="Supplies & Book Tracking",
        start=None, end=None,
        status="Blocked", who="VizBytes Engineering", change="Awaiting the school",
        notes="Awaiting the per-level supplies and books list.",
        bar=None)),
    ("act", dict(
        activity="Scheme of Work",
        start=None, end=None,
        status="Open", who="VizBytes Engineering", change="Reopened 21 Aug",
        notes="Reopened by Ms Christina so a substitute can see lesson topics. "
              "Proposal: a link per class to where the scheme already lives, "
              "not a rebuild - it was built and removed twice before.",
        bar=None)),
    ("act", dict(
        activity="Transcript of Records (TOR)",
        start=None, end=None,
        status="Blocked", who="VizBytes Engineering", change="Awaiting Ms Wynne",
        notes="The underlying data is ready. Awaiting the school's TOR "
              "template; temporary and permanent samples received.",
        bar=None)),
    ("act", dict(
        activity="AEB Grade Change Request Flow",
        start=None, end=None,
        status="Blocked", who="VizBytes Engineering",
        change="Awaiting Ms Chandana",
        notes="Approval order agreed 21 Aug (sequential). Awaiting the AEB "
              "approval form and confirmation of who sits on the board.",
        bar=None)),
]


# --------------------------------------------------------------------------
# Delivery log - the evidence behind every bar on the Timeline sheet.
# --------------------------------------------------------------------------

DELIVERY_LOG = [
    ("2026-03-17", "Project kickoff and scope confirmation", "Original plan sheet (plan-dated)"),
    ("2026-03-23", "User research and requirements gathering", "Original plan sheet (plan-dated)"),
    ("2026-04-14", "Repository created; foundation, roster, grade entry, locking and audit trail", "First commit; migrations 001-007"),
    ("2026-04-14", "Teacher assignments", "Migration 003"),
    ("2026-04-15", "Grade change-request workflow", "Migration 009"),
    ("2026-04-16", "Live change-request badge; report card generation and design pass", "Migration 010"),
    ("2026-04-17", "P-Files document repository (pivoted from a review queue the same day)", "Migration 011"),
    ("2026-04-17", "Admissions dashboard", "Sprint 7"),
    ("2026-04-17", "Records module phases 1-2; report card split into interim and final", "Sprint 10"),
    ("2026-04-18", "Records module phase 3", "Sprint 10"),
    ("2026-04-20", "SIS Admin hub, year setup wizard, module switcher, approver routing", "Migrations 012-013"),
    ("2026-04-21", "Attendance module - daily register and calendar", "Migrations 014-015"),
    ("2026-04-22", "Student Evaluation module; school calendar with five day types", "Migrations 018-019"),
    ("2026-04-23", "Admissions split into its own module", "Migrations 023-026"),
    ("2026-04-24", "Dashboard framework and Insights layer", "Sprint 21"),
    ("2026-04-24", "DROPPED: SharePoint inquiry sync, inquiry tracking, SIS analytics phase", "Reverted"),
    ("2026-04-25", "Drill-down framework across four modules", "Sprint 22"),
    ("2026-04-27", "Global command palette; structured residence history", "Sprint 27"),
    ("2026-04-29", "Parent portal sign-in and report card access; P-Files renewal lifecycle", "Migrations 030-034"),
    ("2026-04-30", "Mid-year section transfer", "Sprint 30"),
    ("2026-05-05", "Demo dataset built", "Seeder"),
    ("2026-05-06", "Compare mode across six modules", "Sprint 34"),
    ("2026-05-08", "Parent publication and P-Files reminder emails", "Sprint 35"),
    ("2026-05-12", "Unified data tables; direct user creation", "Sprint 37"),
    ("2026-05-13", "Change-request hardening; per-topic evaluation ratings", "Migrations 044-047"),
    ("2026-05-14", "Masterfile with subject and overall academic awards; vacation-leave quotas", "Migrations 048-049"),
    ("2026-05-14", "EXECUTIVE DEMO", "Milestone"),
    ("2026-05-19", "Non-examinable subject score entry; school letterhead; STP application status", "Migrations 050-056"),
    ("2026-05-20", "Assessment slot metadata; Scheme of Work builder", "Migrations 057-058"),
    ("2026-05-25", "Scheme of Work flipped to teacher-owned", "Migrations 061-063"),
    ("2026-05-28", "WITHDRAWN: Scheme of Work removed entirely", "Migrations 065-066"),
    ("2026-05-28", "Automated test suite and release pipeline established (73 tests)", "Sprint 43"),
    ("2026-05-29", "Attendance proration for late enrollees", "Migrations 067-068"),
    ("2026-05-30", "Evaluation scope narrowed to adviser write-ups only", "Sprint 47"),
    ("2026-06-01", "Attendance daily view; student lookup sheet", "Sprint 49"),
    ("2026-06-04", "School calendar rebuilt; plain-English audit log; Masterfile Excel export", "Sprint 54"),
    ("2026-06-05", "One-click approve/reject for grade change requests from email", "Sprint 55"),
    ("2026-06-06", "Academic Summary hub; section index numbers", "Migrations 071-072"),
    ("2026-06-10", "Insights dashboards for all four modules", "Sprint 59"),
    ("2026-06-11", "Structured section schedule", "Migration 074"),
    ("2026-06-15", "Client data layer rebuilt across all 80 client screens", "Sprint 61"),
    ("2026-06-23", "Insights two-year comparison across all dashboards", "Sprint 62"),
    ("2026-06-26", "Attendance term sheet reproduced in Excel, with export", "Sprint 63"),
    ("2026-06-26", "Year Setup Workbench", "Sprint 63"),
    ("2026-07-07", "Admissions nine-stage pipeline strip; codebase bug hunt", "Sprint 64"),
    ("2026-07-09", "AY2025 historical grade backfill (Terms 1-3)", "Backfill scripts"),
    ("2026-07-10", "AY2025 backfill complete; Term 4 blocked - no source workbooks", "Backfill scripts"),
    ("2026-07-11", "Grade Levels and progression module", "Migration 078"),
    ("2026-07-14", "SIS Admin navigation redesign", "Sprint 67"),
    ("2026-07-16", "Unified subject setup and weights", "Migrations 079-085"),
    ("2026-07-17", "AY2026 backfill begins - enrolment, attendance", "Backfill scripts"),
    ("2026-07-20", "AY2026 Term 1 and 2 backfill complete (grades, evaluation)", "Backfill scripts"),
    ("2026-07-22", "Insights rebuilt in the current visual language", "Sprint 67"),
    ("2026-07-23", "Role rename: registrar to academic coordinator", "Migration 092"),
    ("2026-07-25", "Home page and account page rebuilt", "Sprint 67"),
    ("2026-07-28", "Notification bell in every module; sensitive-field masking", "Sprint 67"),
    ("2026-07-29", "Classroom module", "Migration 094"),
    ("2026-07-30", "Permissions layer - editable grants per role; demo environment retired", "Migrations 101-104"),
    ("2026-07-31", "ACADEMICS TRAINING SESSION 1 HELD - 11 feedback items raised", "Milestone"),
    ("2026-08-03", "Excused-absence reason; House colours; whole-year trend; at-risk components", "Migrations 109-110"),
    ("2026-08-05", "Enrolment and class assignment separated into two steps", "Sprint 67"),
    ("2026-08-06", "Four houses named; student house allocation imported", "Migration 111"),
    ("2026-08-09", "Teacher-facing student drawer; adviser's at-risk list", "Sprint 67"),
    ("2026-08-12", "Relief teachers, first version", "Migrations 112-116"),
    ("2026-08-13", "ADMIN TRAINING SESSION 1 HELD - 13 action items raised", "Milestone"),
    ("2026-08-13", "Relief simplified to a single cover field; one subject teacher per class", "Migrations 117-118"),
    ("2026-08-14", "Class assignment defects fixed; profile and family sheet saving fixed", "Migration 119"),
    ("2026-08-17", "Save/confirm feedback complete across all 80 write screens", "Sprint 67"),
    ("2026-08-17", "School-wide academic summary; nationality insights and export", "Sprint 67"),
    ("2026-08-18", "Awards made school-wide; disciplinary records schema applied", "Migrations 120-122"),
    ("2026-08-21", "Disciplinary records - five screens and a school-wide register", "Migrations 120-122"),
    ("2026-08-21", "Student lookup rebuilt - whole roster, whole year", "Sprint 67"),
    ("2026-08-24", "One measure at a time on the lookup panel; two chart defects fixed", "Sprint 67"),
    ("2026-08-24", "Relief cover start and end dates - NOT YET APPLIED OR VERIFIED", "Migration 123"),
]


# --------------------------------------------------------------------------
# Rendering
# --------------------------------------------------------------------------

def thin(color=GRIDLINE):
    return Side(style="thin", color=color)


def cell_border(col_offset, is_today, is_month_start):
    left = thin()
    if is_today:
        left = Side(style="medium", color=RED)
    elif is_month_start:
        left = Side(style="thin", color="8EA9DB")
    return Border(left=left, right=thin(), top=thin(), bottom=thin())


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Timeline"

    n_weeks = len(WEEKS)
    last_col = FIRST_WEEK_COL + n_weeks - 1
    today = date.today()
    today_col = FIRST_WEEK_COL + week_index(today)
    month_start_cols = set()
    seen_months = set()
    for i, w in enumerate(WEEKS):
        key = (w.year, w.month)
        if key not in seen_months:
            seen_months.add(key)
            month_start_cols.add(FIRST_WEEK_COL + i)

    # ---- column widths ----
    widths = {1: 5, 2: 46, 3: 24, 4: 11, 5: 20, 6: 22, 7: 18, 8: 62}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for i in range(n_weeks):
        ws.column_dimensions[get_column_letter(FIRST_WEEK_COL + i)].width = 3.1

    # ---- row 1: title band ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    t = ws.cell(row=1, column=1, value="IMPLEMENTATION TIMELINE  |  HFSE INTERNATIONAL SCHOOL")
    t.font = Font(name="Calibri", size=16, bold=True, color=CYAN_TEXT)
    t.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, last_col + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 30

    # ---- row 2: subtitle ----
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    s = ws.cell(row=2, column=1,
                value="Student Information System (SIS) - Actual Delivery Against Plan  |  "
                      "17 Mar - 25 Aug 2026  |  Phase 2 backlog carried forward")
    s.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, last_col + 1):
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[2].height = 20

    # ---- row 3: legend + stamp ----
    stamp = ws.cell(row=3, column=2,
                    value=f"Generated {today.strftime('%d %b %Y')} from the project "
                          f"repository. Evidence for every bar is on the 'Delivery Log' sheet.")
    stamp.font = Font(name="Calibri", size=9, italic=True, color="595959")
    stamp.alignment = Alignment(horizontal="left", vertical="center")

    legend = [(GREEN, "Delivered"), (AMBER, "In progress"),
              (GREY, "Withdrawn"), (LIGHTBLUE, "Planned")]
    lc = FIRST_WEEK_COL
    for color, label in legend:
        ws.merge_cells(start_row=3, start_column=lc, end_row=3, end_column=lc + 1)
        for k in range(2):
            ws.cell(row=3, column=lc + k).fill = PatternFill("solid", fgColor=color)
            ws.cell(row=3, column=lc + k).border = Border(
                left=thin("808080"), right=thin("808080"),
                top=thin("808080"), bottom=thin("808080"))
        ws.merge_cells(start_row=3, start_column=lc + 2, end_row=3, end_column=lc + 5)
        lab = ws.cell(row=3, column=lc + 2, value=label)
        lab.font = Font(name="Calibri", size=8)
        lab.alignment = Alignment(horizontal="left", vertical="center")
        lc += 7
    ws.row_dimensions[3].height = 16

    # ---- row 4: spacer ----
    ws.row_dimensions[4].height = 6

    # ---- row 5: month band ----
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=8)
    pt = ws.cell(row=5, column=1, value="PROJECT TIMELINE")
    pt.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    pt.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, 9):
        ws.cell(row=5, column=c).fill = PatternFill("solid", fgColor=BLUE)

    run_start, run_key = FIRST_WEEK_COL, (WEEKS[0].year, WEEKS[0].month)
    for i in range(n_weeks + 1):
        key = (WEEKS[i].year, WEEKS[i].month) if i < n_weeks else None
        if key != run_key:
            end = FIRST_WEEK_COL + i - 1
            ws.merge_cells(start_row=5, start_column=run_start, end_row=5, end_column=end)
            m = ws.cell(row=5, column=run_start, value=MONTHS[run_key[1] - 1].upper())
            m.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
            m.alignment = Alignment(horizontal="center", vertical="center")
            for c in range(run_start, end + 1):
                ws.cell(row=5, column=c).fill = PatternFill("solid", fgColor=BLUE)
                ws.cell(row=5, column=c).border = Border(
                    left=thin(WHITE), right=thin(WHITE))
            run_start, run_key = FIRST_WEEK_COL + i, key
    ws.row_dimensions[5].height = 18

    # ---- row 6: header ----
    headers = ["No.", "Activity", "Dates", "Duration", "Status",
               "Responsible", "Change", "Notes"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin(WHITE), right=thin(WHITE),
                             top=thin(WHITE), bottom=thin(WHITE))
    for i, w in enumerate(WEEKS):
        cell = ws.cell(row=6, column=FIRST_WEEK_COL + i, value=w.day)
        cell.font = Font(name="Calibri", size=7, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin(WHITE), right=thin(WHITE),
                             top=thin(WHITE), bottom=thin(WHITE))
    ws.row_dimensions[6].height = 26

    # ---- body ----
    r = 7
    n = 0
    zebra = False
    for kind, payload in ROWS:
        if kind == "band":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            b = ws.cell(row=r, column=1, value=payload)
            b.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for c in range(1, last_col + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BLUE)
            ws.row_dimensions[r].height = 18
            r += 1
            zebra = False
            continue

        n += 1
        d = payload
        start, end = d.get("start"), d.get("end")
        bg = ROW_ALT if zebra else WHITE
        zebra = not zebra

        values = [
            n,
            d["activity"],
            d.get("dates") or fmt_dates(start, end),
            fmt_duration(start, end),
            d["status"],
            d["who"],
            d["change"],
            d["notes"],
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = Border(left=thin(), right=thin(), top=thin(), bottom=thin())
            if c == 1:
                cell.font = Font(name="Calibri", size=9, color="7F7F7F")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 2:
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True, indent=1)
            elif c == 8:
                cell.font = Font(name="Calibri", size=8.5, color="404040")
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True, indent=1)
            elif c == 5:
                fillc = STATUS_FILLS.get(d["status"])
                if fillc:
                    cell.fill = PatternFill("solid", fgColor=fillc)
                    cell.font = Font(name="Calibri", size=9, bold=True,
                                     color=WHITE if fillc in (GREY,) else BLACK)
                else:
                    cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
            else:
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)

        # week cells + bar
        bar = d.get("bar")
        bi = ei = None
        if bar and start and end:
            bi, ei = week_index(start), week_index(end)
        for i in range(n_weeks):
            col = FIRST_WEEK_COL + i
            cell = ws.cell(row=r, column=col)
            if bi is not None and bi <= i <= ei:
                cell.fill = PatternFill("solid", fgColor=BAR_FILLS[bar])
            else:
                cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = cell_border(i, col == today_col, col in month_start_cols)

        ws.row_dimensions[r].height = 30
        r += 1

    # ---- footer band ----
    last_row = r
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=last_col)
    f = ws.cell(row=last_row, column=1,
                value=f"Total: 17 Mar - {today.strftime('%d %b %Y')}  |  "
                      f"{n} activities  |  9 phases delivered + extended build  |  "
                      f"Phase 2: 3 of 10 shipped, 7 awaiting the school")
    f.font = Font(name="Calibri", size=11, bold=True, color=CYAN_TEXT)
    f.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, last_col + 1):
        ws.cell(row=last_row, column=c).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[last_row].height = 26

    # ---- notes under the chart ----
    nr = last_row + 2
    notes = [
        "HOW TO READ THIS SHEET",
        "| Bars run across weeks. Each column is one week, labelled with the date of its Monday. "
        "The red vertical line is today.",
        "| The 'Change' column says how each row differs from the March plan: As planned / Date corrected / "
        "Added / Cancelled.",
        "| Phase 7b lists work that was delivered but never appeared on the original plan - it is the main "
        "reason the project ran past June.",
        "| Rows with no bar have no agreed date. They are shown as blocked rather than given a date the "
        "school has not committed to.",
        "| Dates before 14 Apr 2026 come from the original plan sheet; the code repository begins on that date. "
        "Everything after is evidence-dated.",
        "| Migration 123 (relief cover dates) is written but not yet applied to production, and relief cover has "
        "not been checked in the browser.",
    ]
    for i, text in enumerate(notes):
        ws.merge_cells(start_row=nr + i, start_column=2, end_row=nr + i, end_column=8)
        c = ws.cell(row=nr + i, column=2, value=text)
        if i == 0:
            c.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        else:
            c.font = Font(name="Calibri", size=9, color="404040")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[nr + i].height = 15

    # ---- view + print ----
    ws.freeze_panes = ws.cell(row=7, column=FIRST_WEEK_COL)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "5:6"

    # ---- sheet 2: delivery log ----
    ws2 = wb.create_sheet("Delivery Log")
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 86
    ws2.column_dimensions["C"].width = 36

    ws2.merge_cells("A1:C1")
    h = ws2.cell(row=1, column=1, value="DELIVERY LOG  -  evidence behind every bar on the Timeline sheet")
    h.font = Font(name="Calibri", size=13, bold=True, color=CYAN_TEXT)
    h.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 4):
        ws2.cell(row=1, column=c).fill = PatternFill("solid", fgColor=NAVY)
    ws2.row_dimensions[1].height = 26

    for c, label in enumerate(["Date", "Delivered", "Evidence"], start=1):
        cell = ws2.cell(row=2, column=c, value=label)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for i, (d, what, ev) in enumerate(DELIVERY_LOG):
        rr = 3 + i
        bg = ROW_ALT if i % 2 else WHITE
        milestone = ev == "Milestone"
        for c, v in enumerate([d, what, ev], start=1):
            cell = ws2.cell(row=rr, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = Border(left=thin(), right=thin(), top=thin(), bottom=thin())
            cell.font = Font(name="Calibri", size=9,
                             bold=milestone,
                             color=NAVY if milestone else BLACK)
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=(c == 2), indent=1)

    ws2.freeze_panes = "A3"
    ws2.sheet_view.showGridLines = False

    return wb, n


def main():
    out = Path(__file__).resolve().parent.parent / "HFSE-SIS-Implementation-Timeline.xlsx"
    wb, n = build()
    try:
        wb.save(out)
    except PermissionError:
        sys.stderr.write(
            f"Could not write {out.name} - it is probably open in Excel. "
            f"Close it and run again.\n"
        )
        return 1
    print(f"Wrote {out}")
    print(f"  Timeline sheet:     {n} activities across {len(WEEKS)} weekly columns "
          f"({GRID_START} to {GRID_END})")
    print(f"  Delivery Log sheet: {len(DELIVERY_LOG)} dated entries")
    return 0


if __name__ == "__main__":
    sys.exit(main())
