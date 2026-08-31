# scripts/update-proposal-timeline.py
#
# Updates HFSE-SIS-Proposal-May2026-v3 (2).xlsx.
#
# The proposal is already styled. This script does not restyle it - every new
# cell takes its style by copying an existing cell from the same sheet.
#
# What changes on "Implementation Timeline":
#   * the single text "Dates" column is split into real Start / End date cells
#     (a column is inserted, so Duration..Notes shift one right)
#   * Duration becomes a formula, End - Start + 1, so it can never drift
#   * dates, statuses and notes are updated to what actually happened
#   * row structure is untouched - same 24 activities, same phases, no inserts
#
# "Gantt Chart" is added as a new sheet, styled from the timeline sheet's own
# header, band and body cells.
#
# Run:  python scripts/update-proposal-timeline.py

import shutil
import sys
from copy import copy
from datetime import date, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("Needs openpyxl:  pip install openpyxl\n")
    sys.exit(1)

BOOK = "HFSE-SIS-Proposal-May2026-v3 (2).xlsx"
SHEET = "Implementation Timeline"
GANTT = "Gantt Chart"

D = date
TODAY = D(2026, 8, 28)

# Column layout AFTER the insert:
#   B No. | C Activity | D Start | E End | F Duration | G Status | H Who | I Notes
NEW_LAST = 9                     # column I
DATE_FMT = "d mmm yyyy"
DUR_FMT = '0" days";;""'

# row -> (start, end, status, notes). None dates mean the row has no date yet;
# an end of None with a real start means the work is still running.
PLAN = {
    9:  (D(2026, 3, 17), D(2026, 3, 22), "Finished", None),
    11: (D(2026, 3, 23), D(2026, 3, 29), "Finished", None),
    13: (D(2026, 4, 14), D(2026, 7, 14), "Finished", None),
    15: (D(2026, 4, 14), D(2026, 6, 6),  "Finished", None),
    16: (D(2026, 4, 17), D(2026, 6, 12), "Finished", None),
    17: (D(2026, 4, 17), D(2026, 8, 5),  "Finished", None),
    18: (D(2026, 4, 17), D(2026, 7, 30), "Finished", None),
    20: (D(2026, 4, 21), D(2026, 6, 26), "Finished", None),
    21: (D(2026, 4, 22), D(2026, 5, 30), "Finished", None),
    22: (D(2026, 4, 14), D(2026, 8, 28), "Finished",
         "Continuous; 3,152 automated checks green"),
    24: (D(2026, 5, 5),  D(2026, 7, 30), "Finished",
         "Retired 30 Jul once real school data loaded"),
    25: (D(2026, 5, 14), D(2026, 5, 18), "Finished", None),
    27: (D(2026, 5, 18), D(2026, 8, 28), "Finished",
         "Ran to 28 Aug, not 5 Jul"),
    28: (D(2026, 7, 9),  D(2026, 7, 20), "Finished",
         "AY2025 and AY2026 past terms loaded"),
    30: (D(2026, 7, 15), D(2026, 7, 19), "Finished", None),
    32: (None, None, "CANCELLED", "Training moved to virtual"),
    33: (D(2026, 7, 31), D(2026, 7, 31), "Finished",
         "Held 2 weeks early; 11 feedback items"),
    34: (D(2026, 8, 13), D(2026, 8, 13), "Finished",
         "Held early; 13 action items raised"),
    35: (None, None, "On hold", "19 Aug passed; new date needed"),
    36: (D(2026, 8, 30), D(2026, 8, 30), "Upcoming", None),
    38: (D(2026, 8, 3),  None, "Ongoing",
         "2 of 10 in use, 2 built, 6 awaiting school input"),
    39: (D(2026, 8, 13), None, "Ongoing",
         "Alongside the current sheets since 13 Aug"),
    40: (None, None, "Not set", "Follows a clean parallel run"),
    41: (D(2026, 8, 25), None, "Ongoing", "Hypercare period"),
}
RENAMES = {28: "Historical Data Migration"}

SUBTITLE = ("Student Information System (SIS) – Actual Delivery Against Plan "
            "(Mar 17 – Aug 28, 2026)")
FOOTER = ("Total Implementation Period: 17 March – 28 August 2026 "
          "(~24 weeks, 9 phases) + post-launch support, ongoing since 25 August")

PHASE_ROWS = [8, 10, 12, 14, 19, 23, 26, 29, 31, 37]
GRID_START, GRID_END = D(2026, 3, 16), D(2026, 10, 26)
MON = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
       "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]


def weeks():
    out, cur = [], GRID_START
    while cur <= GRID_END:
        out.append(cur)
        cur += timedelta(days=7)
    return out


WEEKS = weeks()


def restyle(target, source):
    """Give `target` the exact style of `source` - nothing invented."""
    target._style = copy(source._style)


def split_dates_column(ws):
    """Insert a column so Dates becomes Start + End. Merges are rebuilt by hand
    because openpyxl does not move them for us."""
    old = sorted(str(m) for m in ws.merged_cells.ranges)
    for m in old:
        ws.unmerge_cells(m)

    ws.insert_cols(5)                                   # new blank column E

    for row in range(7, 43):                            # E inherits D's styling
        restyle(ws.cell(row=row, column=5), ws.cell(row=row, column=4))

    for m in old:                                       # ...:H -> ...:I
        ws.merge_cells(m.replace("H", get_column_letter(NEW_LAST)))

    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws["D7"], ws["E7"] = "Start", "End"


def update_table(ws):
    split_dates_column(ws)
    finished, plain = ws["G9"], ws["G32"]               # status styles, post-shift

    for row, (start, end, status, notes) in PLAN.items():
        d, e, f = ws.cell(row=row, column=4), ws.cell(row=row, column=5), \
                  ws.cell(row=row, column=6)
        d.value, d.number_format = start, DATE_FMT
        e.value, e.number_format = end, DATE_FMT
        f.value = f"=IF(OR(D{row}=\"\",E{row}=\"\"),\"\",E{row}-D{row}+1)"
        f.number_format = DUR_FMT

        g = ws.cell(row=row, column=7)
        g.value = status
        restyle(g, finished if status == "Finished" else plain)

        if notes:
            ws.cell(row=row, column=9).value = notes
        if row in RENAMES:
            ws.cell(row=row, column=3).value = RENAMES[row]

    ws["B4"].value = SUBTITLE
    ws["B42"].value = FOOTER
    return len(PLAN)


def build_gantt(wb):
    """A Gantt whose Activity / Start / End / Duration cells are LINKED to the
    timeline sheet, and whose bars are conditional formatting driven by those
    linked dates - so editing a date on the timeline moves the bar here."""
    src = wb[SHEET]
    if GANTT in wb.sheetnames:
        del wb[GANTT]
    ws = wb.create_sheet(GANTT, wb.sheetnames.index(SHEET) + 1)
    ref = f"'{SHEET}'"

    title_s, sub_s = src["B3"], src["B4"]
    head_s, band_s = src["B7"], src["B8"]
    body_s, name_s = src["F9"], src["C9"]
    date_s, foot_s = src["D9"], src["B42"]
    green = src["G9"].fill.start_color.rgb          # the sheet's own green

    first = 6                                        # grid starts at column F
    last = first + len(WEEKS) - 1
    ws.column_dimensions["A"].width = src.column_dimensions["A"].width
    ws.column_dimensions["B"].width = src.column_dimensions["C"].width
    for col, source in (("C", "D"), ("D", "E"), ("E", "F")):
        ws.column_dimensions[col].width = src.column_dimensions[source].width
    for i in range(len(WEEKS)):
        ws.column_dimensions[get_column_letter(first + i)].width = 2.6

    ws["B2"] = "GANTT CHART | HFSE INTERNATIONAL SCHOOL"
    restyle(ws["B2"], title_s)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=16)
    ws["B3"] = (f"Linked to the {SHEET} sheet - edit a date there and the bar "
                f"below moves")
    restyle(ws["B3"], sub_s)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=16)
    ws.row_dimensions[2].height = src.row_dimensions[3].height
    ws.row_dimensions[3].height = src.row_dimensions[4].height

    for col, label in ((2, "Activity"), (3, "Start"), (4, "End"), (5, "Duration")):
        ws.cell(row=5, column=col, value=label)
        restyle(ws.cell(row=5, column=col), head_s)
        restyle(ws.cell(row=6, column=col), head_s)
        ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)

    run, key = first, (WEEKS[0].year, WEEKS[0].month)
    for i in range(len(WEEKS) + 1):
        k = (WEEKS[i].year, WEEKS[i].month) if i < len(WEEKS) else None
        if k != key:
            end = first + i - 1
            for c in range(run, end + 1):
                restyle(ws.cell(row=5, column=c), head_s)
            ws.cell(row=5, column=run, value=MON[key[1] - 1])
            ws.merge_cells(start_row=5, start_column=run, end_row=5, end_column=end)
            run, key = first + i, k
    for i, w in enumerate(WEEKS):
        cell = ws.cell(row=6, column=first + i, value=w)   # a REAL date, shown as the day
        restyle(cell, head_s)
        cell.number_format = "d"
    ws.row_dimensions[5].height = src.row_dimensions[7].height
    ws.row_dimensions[6].height = src.row_dimensions[7].height

    out = 7
    for row in sorted(PLAN) + [None]:
        for band in [b for b in PHASE_ROWS if row and b == row - 1]:
            for c in range(2, last + 1):
                restyle(ws.cell(row=out, column=c), band_s)
            ws.cell(row=out, column=2, value=f"={ref}!B{band}")
            ws.merge_cells(start_row=out, start_column=2, end_row=out, end_column=5)
            ws.row_dimensions[out].height = src.row_dimensions[band].height
            out += 1
        if row is None:
            break

        # every one of these four is a live link back to the timeline
        for col, src_col, style in ((2, "C", name_s), (3, "D", date_s),
                                    (4, "E", date_s), (5, "F", body_s)):
            cell = ws.cell(row=out, column=col, value=f"={ref}!{src_col}{row}")
            restyle(cell, style)
            if src_col in ("D", "E"):
                cell.number_format = DATE_FMT
            if src_col == "F":
                cell.number_format = DUR_FMT

        for i in range(len(WEEKS)):
            restyle(ws.cell(row=out, column=first + i), body_s)
        ws.row_dimensions[out].height = src.row_dimensions[row].height
        out += 1

    # one rule for the whole grid: light the week if it overlaps Start..End.
    # A blank End means the work is still running, so the bar reaches today.
    grid = f"{get_column_letter(first)}7:{get_column_letter(last)}{out - 1}"
    rule = FormulaRule(
        formula=[f'AND($C7<>"",{get_column_letter(first)}$6+6>=$C7,'
                 f'{get_column_letter(first)}$6<=IF($D7="",TODAY(),$D7))'],
        fill=PatternFill("solid", start_color=green, end_color=green),
        stopIfTrue=False,
    )
    ws.conditional_formatting.add(grid, rule)

    out += 1
    for c in range(2, last + 1):
        restyle(ws.cell(row=out, column=c), foot_s)
    ws.cell(row=out, column=2,
            value="Bars are drawn from the Start and End dates on the "
                  "Implementation Timeline sheet. A row with no End date is "
                  "still running and its bar reaches today.")
    ws.merge_cells(start_row=out, start_column=2, end_row=out, end_column=last)

    ws.freeze_panes = ws.cell(row=7, column=first)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return out


def main():
    root = Path(__file__).resolve().parent.parent
    src = root / BOOK
    bak = src.with_suffix(".bak.xlsx")
    if not src.exists():
        sys.stderr.write(f"Not found: {src}\n")
        return 1

    if bak.exists():
        shutil.copy2(bak, src)          # always start from the pristine original
    else:
        shutil.copy2(src, bak)

    try:
        wb = load_workbook(src)
        rows = update_table(wb[SHEET])
        cols = build_gantt(wb)
        wb.save(src)
    except PermissionError:
        sys.stderr.write("The workbook is open in Excel. Close it and re-run.\n")
        return 1

    print(f"Updated {src.name}   (original preserved as {bak.name})")
    print(f"  '{SHEET}': Dates split into Start/End, Duration is now a formula")
    print(f"  {rows} activity rows updated; styling copied from the sheet itself")
    print(f"  '{GANTT}': new sheet, {len(WEEKS)} weekly columns, {cols} rows")
    print(f"  untouched: {[s for s in wb.sheetnames if s not in (SHEET, GANTT)]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
